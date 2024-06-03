import os
import django
import openpyxl
from django.db import transaction, connection

# Django 프로젝트 설정을 불러옴
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from ai_chatbot.models import ChatbotTrainData  # 모델이 정의된 앱명으로 변경
# 학습 데이터 초기화
def all_clear_train_data():
    ChatbotTrainData.objects.all().delete()
    # auto increment 초기화
    # with connection.cursor() as cursor:
    #     cursor.execute("DELETE FROM sqlite_sequence WHERE name='ChatbotTrainData';")


# db에 데이터 저장
def insert_data(xls_row):
    intent, ner, query, answer, answer_img_url = xls_row

    # 엑셀에서 불러온 cell에 데이터가 없는 경우, None으로 치환
    ChatbotTrainData.objects.create(
        intent=intent.value if intent.value else None,
        ner=ner.value if ner.value else None,
        query=query.value if query.value else None,
        answer=answer.value if answer.value else None,
        answer_image=answer_img_url.value if answer_img_url.value else None
    )


train_file = './train_data_최종보완.xlsx'

try:
    # 기존 학습 데이터 초기화
    all_clear_train_data()

    # 학습 엑셀 파일 불러오기
    wb = openpyxl.load_workbook(train_file)
    sheet = wb['train_data_최종']  # 시트명 확인 잘 하기

    with transaction.atomic():  # 트랜잭션 관리
        for row in sheet.iter_rows(min_row=2):  # 헤더는 불러오지 않음
            # 데이터 저장
            insert_data(row)  # too many values to unpack (expected 5) => .xlsx파일에 빈 열이 생기지 않았는지 확인

    wb.close()

except Exception as e:
    print(e)




# import pymysql
# import openpyxl
#
# from config.DatabaseConfig import * # DB 접속 정보 불러오기
#
#
# # 학습 데이터 초기화
# def all_clear_train_data(db):
#     # 기존 학습 데이터 삭제
#     sql = '''
#             delete from chatbot_train_data
#         '''
#     with db.cursor() as cursor:
#         cursor.execute(sql)
#
#     # auto increment 초기화
#     sql = '''
#     ALTER TABLE chatbot_train_data AUTO_INCREMENT=1
#     '''
#     with db.cursor() as cursor:
#         cursor.execute(sql)
#
#
# # db에 데이터 저장
# def insert_data(db, xls_row):
#     intent, ner, query, answer, answer_img_url = xls_row
#
#     sql = '''
#         INSERT chatbot_train_data(intent, ner, query, answer, answer_image)
#         values(
#          '%s', '%s', '%s', '%s', '%s'
#         )
#     ''' % (intent.value, ner.value, query.value, answer.value, answer_img_url.value)
#
#     # 엑셀에서 불러온 cell에 데이터가 없는 경우, null 로 치환
#     sql = sql.replace("'None'", "null")
#
#     with db.cursor() as cursor:
#         cursor.execute(sql)
#         #print('{} 저장'.format(query.value))
#         db.commit()
#
#
# train_file = './ai_chatbot/train_tools/qna/train_data_최종1.xlsx'
# db = None
# try:
#     # DB 호스트 정보에 맞게 입력해주세요
#     db = pymysql.connect(
#         host='127.0.0.1',
#         user='root',
#         db='chatbot',
#         charset='utf8'
#     )
#
#     # 기존 학습 데이터 초기화
#     all_clear_train_data(db)
#
#     # 학습 엑셀 파일 불러오기
#     wb = openpyxl.load_workbook(train_file)
#     sheet = wb['train_data_최종'] # 시트명 확인 잘 하기
#     for row in sheet.iter_rows(min_row=2): # 해더는 불러오지 않음
#         # 데이터 저장
#         insert_data(db, row) # too many values to unpack (expected 5) => .xlsx파일에 빈 열이 생기지 않았는지 확인
#
#     wb.close()
#
# except Exception as e:
#     print(e)
#
# finally:
#     if db is not None:
#         db.close()
