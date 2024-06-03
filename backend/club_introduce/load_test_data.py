import os
import django
import openpyxl
from django.db import transaction

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from club_introduce.models import Club, ClubDetail, ClubMember
club_data_file = '../utils/club_introduce_club.xlsx'
club_detail_data_file = '../utils/club_introduce_clubdetail.xlsx'

def all_clear_train_data():
    Club.objects.all().delete()
    ClubDetail.objects.all().delete()
    ClubMember.objects.all().delete()


def insert_club_data(xls_row):
    club_name, category, introducation, logo, new_club, photo, type = xls_row

    # 엑셀에서 불러온 cell에 데이터가 없는 경우, None으로 치환
    Club.objects.create(
        club_name=club_name.value if club_name.value else None,
        category=category.value if category.value else None,
        type=type.value if type.value else None,
        introducation=introducation.value if introducation.value else None,
        photo=photo.value if photo.value else None,
        logo=logo.value if logo.value else None,
        new_club=new_club.value if new_club.value else False
    )


def insert_club_detail_data(xls_row):
    club_id, join, location, activity, fee = xls_row

    ClubDetail.objects.create(
        # club=club_id.value if club_id.value else None,
        club=Club.objects.get(club_name=club_id.value),
        join=join.value if join.value else '',
        location=location.value if location.value else '',
        activity=activity.value if activity.value else '',
        fee=fee.value if fee.value else 0
    )




try:
    # 기존 학습 데이터 초기화
    # all_clear_train_data()

    # 학습 엑셀 파일 불러오기
    # wb = openpyxl.load_workbook(club_data_file)
    # sheet = wb['Sheet1']  # 시트명 확인 잘 하기
    #
    # with transaction.atomic():  # 트랜잭션 관리
    #     for row in sheet.iter_rows(min_row=2):  # 헤더는 불러오지 않음
    #         # 데이터 저장
    #         insert_club_data(row)  # too many values to unpack (expected 5) => .xlsx파일에 빈 열이 생기지 않았는지 확인
    #
    # wb.close()

    wb = openpyxl.load_workbook(club_detail_data_file)
    sheet = wb['Sheet1']  # 시트명 확인 잘 하기

    with transaction.atomic():  # 트랜잭션 관리
        for row in sheet.iter_rows(min_row=2):  # 헤더는 불러오지 않음
            # 데이터 저장
            insert_club_detail_data(row)  # too many values to unpack (expected 5) => .xlsx파일에 빈 열이 생기지 않았는지 확인

    wb.close()

except Exception as e:
    print(e)